import requests
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

from config_data.config import Config, load_config
from lexicon.city import LEXICON_CITY
from database.database import users_db


config: Config = load_config()
USER_ID = config.tg_bot.user_id
KEY = config.tg_bot.apy_key


def _get_string(s: str) -> str:
    result = s

    if "&" in result:
        result = result.replace("&", "%26")

    return ' '.join(result.split())


def _check_tag(s: str) -> str:
    result = s

    if "%3Cb%3E" in result:
        result = result.replace("%3Cb%3E", " ").replace("%3C/b%3E", " ")

    if "&amp;nbsp," in result:
        result = result.replace("&amp;nbsp,", " ")

    if "&amp;amp;nbsp," in result:
        result = result.replace("&amp;amp;nbsp,", " ")

    if "&lt;hlword&gt;" in result:
        result = result.replace("&lt;hlword&gt;", " ").replace("&lt;/hlword&gt;", " ")

    return result


def get_file(telegram_user_id: int) -> str | None:
    search = users_db[telegram_user_id][0]
    client_string = users_db[telegram_user_id][1]
    client_request = _get_string(client_string)
    region = LEXICON_CITY[users_db[telegram_user_id][2]]
    user_ID = USER_ID
    key = KEY

    name_file = f'{client_request}_{users_db[telegram_user_id][2]}.xlsx'
    wb = Workbook(name_file)
    wb.save(name_file)
    wb = load_workbook(name_file)
    ws = wb.active

    ws.cell(row=1, column=1, value='порядок выдачи')
    ws.cell(row=1, column=2, value='вид')
    ws.cell(row=1, column=3, value='адрес сайта')
    ws.cell(row=1, column=4, value='заголовок')
    ws.cell(row=1, column=5, value='описание')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 15

    wb.save(name_file)

    count_row = 2

    for i in range(2):
        url = f'http://xmlriver.com/search_yandex/xml?user={user_ID}&key={key}&lr={region}&query={client_request}'

        params = {'lang': 'ru',
                  'device': 'desktop',
                  'page': i
                  # 'lr': '37',             # регион поиска согласно яндекс
                  # 'raw': 'page',          # отвечает за получение полного html
                  # 'groupby': 10           # ТОП-10 позиций для сбора
        }

        try:
            response = requests.get(url=url, params=params, timeout=60)  # максимальный таймаут XMLRiver = 60
            soup = BeautifulSoup(response.text, 'xml')
            # print(soup.prettify())
            topads = soup.find('topads')
            bottomads = soup.find('bottomads')
            seo = soup.find('grouping')

            try:
                all_top = topads.find_all('query')
            except Exception:
                pass

            try:
                all_bottom = bottomads.find_all('query')
            except Exception:
                pass

            try:
                all_group = seo.find_all('group')
            except Exception:
                pass

            namber_page = str(soup.find('page'))[-8]

            if search == 1 or search == 3:
                try:
                    count_all = 1
                    for top in all_top:
                        ws.cell(row=count_row, column=1, value=f'СП-{namber_page}-{count_all}')
                        ws.cell(row=count_row, column=2, value=f'Реклама')
                        ws.cell(row=count_row, column=3, value=str(top.find('url'))[5:-6])
                        ws.cell(row=count_row, column=4, value=_check_tag(str(top.find('title'))[7:-8]))
                        ws.cell(row=count_row, column=5, value=_check_tag(str(top.find('snippet'))[9:-10]))
                        ws.cell(row=count_row, column=6, value=' ')
                        if top.find_all('sitelink'):
                            count_column = 0
                            for i in range(len(top.find_all('sitelink'))):
                                ws.cell(row=1, column=6 + count_column, value=f'Быстрая ссылка - {1 + i}')
                                ws.cell(row=1, column=7 + count_column, value=f'Заголовок БС - {1 + i}')
                                ws.cell(row=count_row, column=6 + count_column,
                                        value=str(top.find_all('sitelink')[i].find('url'))[5:-6])
                                ws.cell(row=count_row, column=7 + count_column,
                                        value=_check_tag(str(top.find_all('sitelink')[i].find('title'))[7:-8]))
                                count_column += 2
                        count_row += 1
                        count_all += 1
                        wb.save(name_file)

                except Exception as ex:
                    ws.cell(row=count_row, column=1, value=f'СП-{namber_page}-1')
                    ws.cell(row=count_row, column=2, value=f'Реклама')
                    ws.cell(row=count_row, column=3, value='В данном разделе информации не найдено')
                    ws.cell(row=count_row, column=5, value=f'{ex}')
                    wb.save(name_file)
                    count_row += 1
                    pass

                count_row += 1

                try:
                    count_bottom = 1
                    for bottom in all_bottom:
                        ws.cell(row=count_row, column=1, value=f'M-{namber_page}-{count_bottom}')
                        ws.cell(row=count_row, column=2, value=f'Реклама')
                        ws.cell(row=count_row, column=3, value=str(bottom.find('url'))[5:-6])
                        ws.cell(row=count_row, column=4, value=_check_tag(str(bottom.find('title'))[7:-8]))
                        ws.cell(row=count_row, column=5, value=_check_tag(str(bottom.find('snippet'))[9:-10]))
                        ws.cell(row=count_row, column=6, value=' ')
                        if bottom.find_all('sitelink'):
                            count_column = 0
                            for i in range(len(bottom.find_all('sitelink'))):
                                ws.cell(row=1, column=6 + count_column, value=f'Быстрая ссылка - {1 + i}')
                                ws.cell(row=1, column=7 + count_column, value=f'Заголовок БС - {1 + i}')
                                ws.cell(row=count_row, column=6 + count_column,
                                        value=str(bottom.find_all('sitelink')[i].find('url'))[5:-6])
                                ws.cell(row=count_row, column=7 + count_column,
                                        value=_check_tag(str(bottom.find_all('sitelink')[i].find('title'))[7:-8]))
                                count_column += 2
                        count_row += 1
                        count_bottom += 1
                        wb.save(name_file)

                except Exception as ex:
                    ws.cell(row=count_row, column=1, value=f'M-{namber_page}-1')
                    ws.cell(row=count_row, column=2, value=f'Реклама')
                    ws.cell(row=count_row, column=3, value='В данном разделе информации не найдено')
                    ws.cell(row=count_row, column=5, value=f'{ex}')
                    wb.save(name_file)
                    count_row += 1
                    pass

                count_row += 1

            if search == 2 or search == 3:
                try:
                    count_group = 1
                    for group in all_group:
                        ws.cell(row=count_row, column=1, value=f'{namber_page} - {count_group}')
                        ws.cell(row=count_row, column=2, value=f'SEO')
                        ws.cell(row=count_row, column=3, value=str(group.find('url'))[5:-6])
                        ws.cell(row=count_row, column=4, value=_check_tag(str(group.find('title'))[7:-8]))
                        ws.cell(row=count_row, column=5, value=_check_tag(str(group.find('passage'))[9:-10]))
                        ws.cell(row=count_row, column=6, value=' ')
                        count_row += 1
                        count_group += 1
                        wb.save(name_file)

                except Exception as ex:
                    ws.cell(row=count_row, column=1, value=f'{namber_page}-1')
                    ws.cell(row=count_row, column=2, value=f'SEO')
                    ws.cell(row=count_row, column=3, value='В данном разделе информации не найдено')
                    ws.cell(row=count_row, column=5, value=f'{ex}')
                    wb.save(name_file)
                    count_row += 1
                    pass

                count_row += 1

        except Exception as ex:
            name_file = None
            print(ex)

        time.sleep(1.5)

    return name_file

if __name__ == "__main__":
    get_file()